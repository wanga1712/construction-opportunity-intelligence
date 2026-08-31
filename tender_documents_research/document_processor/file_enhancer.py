from pathlib import Path
from typing import Any, Dict, List, Optional

import docx
from docx.enum.text import WD_COLOR_INDEX
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from utils.logger_config import get_logger


class FileEnhancer:
    def __init__(self) -> None:
        self.logger = get_logger()

    def highlight_matches(self, file_path: Path, matches: List[Dict[str, Any]]) -> bool:
        ext = file_path.suffix.lower()
        if ext == ".docx":
            return self._highlight_docx(file_path, matches)
        if ext in (".xlsx", ".xlsm"):
            return self._highlight_xlsx(file_path, matches)
        return False

    def rename_file(self, file_path: Path, contract_number: str) -> Path:
        contract_safe = self._sanitize_contract(contract_number)
        if not contract_safe:
            return file_path

        # Check if already prefixed
        if file_path.name.startswith(f"{contract_safe}_"):
            return file_path

        new_name = f"{contract_safe}_{file_path.name}"
        new_path = file_path.with_name(new_name)
        try:
            if new_path.exists():
                base = file_path.stem
                suffix = file_path.suffix
                index = 1
                while True:
                    candidate = file_path.with_name(f"{contract_safe}_{base}_{index}{suffix}")
                    if not candidate.exists():
                        new_path = candidate
                        break
                    index += 1
            file_path.rename(new_path)
            return new_path
        except Exception:
            return file_path

    def _highlight_docx(self, file_path: Path, matches: List[Dict[str, Any]]) -> bool:
        try:
            document = docx.Document(str(file_path))
        except Exception:
            return False
        keywords = {
            str(m.get("keyword") or "").lower()
            for m in matches
            if m.get("keyword")
        }
        if not keywords:
            return False
        changed = False
        for paragraph in document.paragraphs:
            text = paragraph.text or ""
            lower = text.lower()
            if not lower:
                continue
            if not any(k in lower for k in keywords):
                continue
            for run in paragraph.runs:
                run_text = run.text or ""
                run_lower = run_text.lower()
                if any(k in run_lower for k in keywords):
                    try:
                        run.font.highlight_color = WD_COLOR_INDEX.YELLOW
                        changed = True
                    except Exception:
                        continue
        if not changed:
            return False
        try:
            document.save(str(file_path))
            return True
        except Exception:
            return False

    def _highlight_xlsx(self, file_path: Path, matches: List[Dict[str, Any]]) -> bool:
        try:
            workbook = load_workbook(filename=str(file_path))
        except Exception:
            return False
        keywords = {
            str(m.get("keyword") or "").lower()
            for m in matches
            if m.get("keyword")
        }
        if not keywords:
            return False
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        changed = False
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if not isinstance(value, str):
                        continue
                    lower = value.lower()
                    if any(k in lower for k in keywords):
                        cell.fill = fill
                        changed = True
        if not changed:
            return False
        try:
            workbook.save(str(file_path))
            return True
        except Exception:
            return False

    def _sanitize_contract(self, value: Optional[str]) -> str:
        if value is None:
            return ""
        result = []
        for ch in str(value):
            if ch.isalnum():
                result.append(ch)
        return "".join(result)

