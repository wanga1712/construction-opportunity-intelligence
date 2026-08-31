from pathlib import Path
from typing import Any, Dict, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class ExcelParser:
    def parse(self, path: Path) -> str:
        text, _ = self.parse_with_meta(path)
        return text

    @staticmethod
    def _read_workbook(
        path: Path, *, data_only: bool
    ) -> Tuple[str, Dict[int, Dict[str, Any]]]:
        workbook = load_workbook(
            filename=str(path), read_only=True, data_only=data_only
        )
        parts: list[str] = []
        line_meta: Dict[int, Dict[str, Any]] = {}
        line_index = 0
        try:
            for sheet in workbook.worksheets:
                sheet_name = sheet.title
                for row in sheet.iter_rows(values_only=False):
                    cell_items: list[dict] = []
                    cell_texts: list[str] = []
                    row_index_val = None
                    for cell in row:
                        value = cell.value
                        if value is None:
                            continue
                        text = str(value).strip()
                        if not text:
                            continue
                        text = text.replace("\n", " ").replace("\r", " ")
                        cell_texts.append(text)
                        cell_items.append({
                            "text": text,
                            "column_letter": get_column_letter(cell.column),
                            "cell_address": cell.coordinate,
                        })
                        row_index_val = cell.row
                    if not cell_texts:
                        continue
                    parts.append(" | ".join(cell_texts))
                    line_index += 1
                    line_meta[line_index] = {
                        "sheet_name": sheet_name,
                        "row_index": row_index_val,
                        "cells": cell_items,
                    }
        finally:
            workbook.close()
        return "\n".join(parts), line_meta

    def parse_with_meta(self, path: Path) -> Tuple[str, Dict[int, Dict[str, Any]]]:
        text, line_meta = self._read_workbook(path, data_only=True)
        if text:
            return text, line_meta
        # Formula-only workbooks often have no cached values. Reopen them with
        # formulas visible instead of incorrectly reporting an empty parse.
        return self._read_workbook(path, data_only=False)
