"""Adapter connecting procurement document parsers to structured table rows."""

from __future__ import annotations

import csv
import io
import os
from pathlib import Path
import re
from typing import Any, Dict, List, Optional, Tuple

from src.product_discovery.dto import ExtractedTableRow


# Regex patterns for detecting column headers
RE_NAME_HDR = re.compile(r"(наименование|название|описание|товар|материал|работа|оборудование|позиция)", re.IGNORECASE)
RE_QTY_HDR = re.compile(r"(кол-?во|колич|объем)", re.IGNORECASE)
RE_UNIT_HDR = re.compile(r"(ед\.\s*изм|единица|измерен)", re.IGNORECASE)
RE_PRICE_HDR = re.compile(r"(цена|стоимость\s+ед|тариф)", re.IGNORECASE)
RE_AMOUNT_HDR = re.compile(r"(сумма|всего|стоимость\s+всего|итого)", re.IGNORECASE)


def parse_numeric_cell(cell_value: Any) -> float:
    """Deterministically extracts clean floating point number from cell value."""
    if cell_value is None:
        return 0.0
    if isinstance(cell_value, (int, float)):
        return float(cell_value)
    
    val_str = str(cell_value).strip().replace("\xa0", " ").replace(" ", "")
    val_str = val_str.replace(",", ".")
    match = re.search(r"[-+]?\d+(?:\.\d+)?", val_str)
    if match:
        try:
            return float(match.group(0))
        except ValueError:
            pass
    return 0.0


def identify_header_mapping(header_cells: List[str]) -> Dict[str, int]:
    """Maps semantic column roles to column indices."""
    mapping: Dict[str, int] = {}
    for idx, cell in enumerate(header_cells):
        text = str(cell).strip()
        if not text:
            continue
        if "name" not in mapping and RE_NAME_HDR.search(text):
            mapping["name"] = idx
        elif "qty" not in mapping and RE_QTY_HDR.search(text):
            mapping["qty"] = idx
        elif "unit" not in mapping and RE_UNIT_HDR.search(text):
            mapping["unit"] = idx
        elif "amount" not in mapping and RE_AMOUNT_HDR.search(text):
            mapping["amount"] = idx
        elif "price" not in mapping and RE_PRICE_HDR.search(text):
            mapping["price"] = idx
    return mapping


class DocumentTableAdapter:
    """Parses Excel, Word, CSV, and text document tables into standardized ExtractedTableRow instances."""

    def __init__(self) -> None:
        pass

    def parse_document_tables(
        self,
        file_path: str | Path,
        procurement_id: int,
        document_id: str = "",
    ) -> List[ExtractedTableRow]:
        """Dispatches file to specific parser and returns list of extracted table rows."""
        p = Path(file_path)
        if not p.exists() or p.stat().st_size == 0:
            return []

        doc_id = document_id or p.stem
        ext = p.suffix.lower()

        if ext in (".xlsx", ".xlsm"):
            return self._parse_xlsx(p, procurement_id, doc_id)
        elif ext == ".csv":
            return self._parse_csv(p, procurement_id, doc_id)
        elif ext in (".docx", ".doc"):
            return self._parse_docx(p, procurement_id, doc_id)
        else:
            return self._parse_plain_text(p, procurement_id, doc_id)

    def _parse_xlsx(self, path: Path, procurement_id: int, doc_id: str) -> List[ExtractedTableRow]:
        """Parses Excel workbook worksheets with header tracking."""
        rows: List[ExtractedTableRow] = []
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
            try:
                for sheet in wb.worksheets:
                    sheet_name = sheet.title
                    current_header_map: Dict[str, int] = {}
                    current_section = "Основной раздел"

                    for r_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                        cell_strs = [str(c).strip() if c is not None else "" for c in row]
                        non_empty = [c for c in cell_strs if c]
                        if not non_empty:
                            continue

                        if len(non_empty) == 1 and len(non_empty[0]) > 5:
                            current_section = non_empty[0]

                        hdr_map = identify_header_mapping(cell_strs)
                        if len(hdr_map) >= 2:
                            current_header_map = hdr_map
                            continue

                        raw_line = " | ".join(non_empty)
                        extracted = ExtractedTableRow(
                            procurement_id=procurement_id,
                            document_id=doc_id,
                            file_path=str(path),
                            sheet_name=sheet_name,
                            page_number=None,
                            table_index=0,
                            row_index=r_idx,
                            raw_cells=cell_strs,
                            raw_text=raw_line,
                            section_name=current_section,
                            column_mapping=dict(current_header_map),
                        )
                        extracted.compute_observation_key()
                        rows.append(extracted)
            finally:
                wb.close()
        except Exception:
            pass
        return rows

    def _parse_csv(self, path: Path, procurement_id: int, doc_id: str) -> List[ExtractedTableRow]:
        """Parses delimited CSV table."""
        rows: List[ExtractedTableRow] = []
        try:
            with open(path, "r", encoding="utf-8", errors="replace") as f:
                first_line = f.readline()
                f.seek(0)
                delimiter = ";" if ";" in first_line else ("\t" if "\t" in first_line else ",")
                reader = csv.reader(f, delimiter=delimiter)
                header_map: Dict[str, int] = {}
                for r_idx, row in enumerate(reader, start=1):
                    cell_strs = [str(c).strip() for c in row]
                    non_empty = [c for c in cell_strs if c]
                    if not non_empty:
                        continue

                    hdr_map = identify_header_mapping(cell_strs)
                    if len(hdr_map) >= 2 and not header_map:
                        header_map = hdr_map
                        continue

                    raw_line = " | ".join(non_empty)
                    extracted = ExtractedTableRow(
                        procurement_id=procurement_id,
                        document_id=doc_id,
                        file_path=str(path),
                        sheet_name="CSV",
                        page_number=None,
                        table_index=0,
                        row_index=r_idx,
                        raw_cells=cell_strs,
                        raw_text=raw_line,
                        section_name="Основной раздел",
                        column_mapping=dict(header_map),
                    )
                    extracted.compute_observation_key()
                    rows.append(extracted)
        except Exception:
            pass
        return rows

    def _parse_docx(self, path: Path, procurement_id: int, doc_id: str) -> List[ExtractedTableRow]:
        """Parses Word docx tables using python-docx."""
        rows: List[ExtractedTableRow] = []
        try:
            import docx
            doc = docx.Document(str(path))
            for t_idx, table in enumerate(doc.tables):
                header_map: Dict[str, int] = {}
                for r_idx, row in enumerate(table.rows, start=1):
                    cell_strs = [c.text.replace("\n", " ").strip() for c in row.cells]
                    non_empty = [c for c in cell_strs if c]
                    if not non_empty:
                        continue

                    hdr_map = identify_header_mapping(cell_strs)
                    if len(hdr_map) >= 2 and not header_map:
                        header_map = hdr_map
                        continue

                    raw_line = " | ".join(non_empty)
                    extracted = ExtractedTableRow(
                        procurement_id=procurement_id,
                        document_id=doc_id,
                        file_path=str(path),
                        sheet_name=f"Table_{t_idx + 1}",
                        page_number=None,
                        table_index=t_idx,
                        row_index=r_idx,
                        raw_cells=cell_strs,
                        raw_text=raw_line,
                        section_name=f"Таблица {t_idx + 1}",
                        column_mapping=dict(header_map),
                    )
                    extracted.compute_observation_key()
                    rows.append(extracted)
        except Exception:
            pass
        return rows

    def _parse_plain_text(self, path: Path, procurement_id: int, doc_id: str) -> List[ExtractedTableRow]:
        """Fallback line parser for plain text / unstructured dumps."""
        rows: List[ExtractedTableRow] = []
        try:
            with open(path, "r", encoding="utf-8", errors="replace") as f:
                for r_idx, line in enumerate(f, start=1):
                    txt = line.strip()
                    if not txt or len(txt) < 3:
                        continue
                    cells = [c.strip() for c in txt.split("|")] if "|" in txt else [txt]
                    extracted = ExtractedTableRow(
                        procurement_id=procurement_id,
                        document_id=doc_id,
                        file_path=str(path),
                        sheet_name="Text",
                        page_number=None,
                        table_index=0,
                        row_index=r_idx,
                        raw_cells=cells,
                        raw_text=txt,
                        section_name="Текст",
                        column_mapping={},
                    )
                    extracted.compute_observation_key()
                    rows.append(extracted)
        except Exception:
            pass
        return rows
